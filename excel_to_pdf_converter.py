# -*- coding: utf-8 -*-
"""
Created on Thu Oct 24 21:49:42 2024

@author: Ken
"""


from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import simpleSplit
from openpyxl.utils import get_column_letter

class ExcelToPDFConverter:
    def __init__(self, font_path):
        # Initialize by registering the Japanese font
        self.register_japanese_font(font_path)

    def register_japanese_font(self, font_path):
        # Register custom Japanese font for use in PDF
        pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))

    def is_border_color_empty(self, color):
        # Check if border color is empty (transparent or white).
        return color is None or (color.type in ['', 'rgb'] and color.rgb.upper() == 'FFFFFF')

    def draw_cell_border(self, c, x, y, width, height):
        # Draw boundary box for cell at given position
        c.setStrokeColorRGB(0, 0, 0)
        c.rect(x, y, width, height, stroke=1, fill=0)

    def is_data_table(self, ws):
        # Determine if worksheet is a complete data table (over 90% cells have data)
        total_cells = ws.max_row * ws.max_column
        non_empty_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
        return (non_empty_cells / total_cells) >= 0.90 if total_cells > 0 else False

    def format_cell_value(self, cell_value):
        # Format cell value: floating-point numbers retain two decimals
        if isinstance(cell_value, float):
            return f"{cell_value:.2f}"
        return str(cell_value)

    def calculate_max_chars(self, cell_width, font_size):
        # Calculate maximum characters in a cell given its width and font size
        char_width = font_size
        return int(cell_width * 1.03 / char_width)

    def adjust_column_widths(self, ws, pdf_width, left_margin, right_margin):
        # Adjust PDF column widths based on Excel column widths
        excel_col_widths = [(ws.column_dimensions[get_column_letter(col)].width or 10) for col in range(1, ws.max_column + 1)]
        total_excel_width = sum(excel_col_widths)
        return [(pdf_width - left_margin - right_margin) * (width / total_excel_width) for width in excel_col_widths]

    def convert_excel_to_pdf(self, excel_file, pdf_file):
        # Convert Excel file to PDF
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        table_flag = self.is_data_table(ws)

        c = canvas.Canvas(pdf_file, pagesize=landscape(A4))
        width, height = landscape(A4)
        left_margin, right_margin, top_margin = 50, 50, 40
        x_offset, y_offset = left_margin, height - top_margin
        base_row_height = 24
        col_widths = self.adjust_column_widths(ws, width, left_margin, right_margin)
        y_current = y_offset
        merged_ranges = ws.merged_cells.ranges

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            cell_heights, row_has_data = [], False

            for cell in row:
                if cell.value is None:
                    continue
                row_has_data = True

                cell_merge_width = col_widths[cell.column - 1]
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        merge_start = merged_range.min_col - 1
                        merge_end = merged_range.max_col
                        cell_merge_width = sum(col_widths[merge_start:merge_end])
                        break

                formatted_value = self.format_cell_value(cell.value)
                lines = simpleSplit(formatted_value, 'JapaneseFont', 12, maxWidth=cell_merge_width)
                line_count = len(lines)
                cell_height = base_row_height * line_count
                cell_heights.append(cell_height)

            if not row_has_data:
                continue

            row_height = max(cell_heights) if cell_heights else base_row_height

            if y_current - row_height < top_margin:
                c.showPage()
                y_current = y_offset

            if table_flag:
                for j, cell in enumerate(row):
                    if cell.value is None:
                        continue

                    merge_width = col_widths[j]
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            merge_start = merged_range.min_col - 1
                            merge_end = merged_range.max_col
                            merge_width = sum(col_widths[merge_start:merge_end])
                            break

                    top_left_x = x_offset + sum(col_widths[:j])
                    top_left_y = y_current
                    
                    self.draw_cell_border(c, top_left_x, top_left_y - row_height, merge_width, row_height)

                    font_size = 12
                    max_chars = self.calculate_max_chars(merge_width, font_size)

                    formatted_value = self.format_cell_value(cell.value)
                    
                    if len(formatted_value) > max_chars:
                        font_size = 10

                    c.setFont('JapaneseFont', font_size)
                    y_text_offset = top_left_y - 15
                    lines = simpleSplit(formatted_value, 'JapaneseFont', font_size, maxWidth=merge_width)
                    for k, line in enumerate(lines):
                        c.drawString(top_left_x + 2, y_text_offset - 12 * k, line)

            else:
                for j, cell in enumerate(row):
                    if cell.value is None:
                        continue

                    merge_width = col_widths[j]
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            merge_start = merged_range.min_col - 1
                            merge_end = merged_range.max_col
                            merge_width = sum(col_widths[merge_start:merge_end])
                            break

                    text = self.format_cell_value(cell.value)

                    font_size = 12
                    max_chars = self.calculate_max_chars(merge_width, font_size)

                    if len(text) > max_chars:
                        needed_width = len(text) * font_size * 0.5
                        remaining_width = merge_width
                        for k in range(j+1, len(row)):
                            if row[k].value is None:
                                remaining_width += col_widths[k]
                                if remaining_width >= needed_width:
                                    break
                            else:
                                break

                        if remaining_width >= needed_width:
                            merge_width = remaining_width
                            max_chars = self.calculate_max_chars(merge_width, font_size)

                    top_left_x = x_offset + sum(col_widths[:j])
                    top_left_y = y_current

                    merge_width = min(merge_width, 61 * 12 * 12 / font_size)
                    lines = text.split('\n')
                    wrapped_lines = []

                    max_chars = int(merge_width * 1.03 / font_size)

                    for line in lines:
                        wrapped_lines.extend([line[i:i+max_chars] for i in range(0, len(line), max_chars)])
                    
                    y_text_offset = top_left_y - 15
                    for k, line in enumerate(wrapped_lines):
                        c.setFont('JapaneseFont', font_size)
                        c.drawString(top_left_x + 2, y_text_offset - 12 * k, line)

            y_current -= row_height
        c.save()
        
        print("Conversion complete: {} to {}".format(excel_file, pdf_file))  # Notification of completion






